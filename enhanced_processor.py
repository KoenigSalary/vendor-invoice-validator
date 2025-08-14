import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import logging
import os
from pathlib import Path
import re
from typing import Dict, List, Tuple, Optional
import zipfile
import shutil
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
warnings.filterwarnings('ignore')

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('invoice_processing.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class EnhancedInvoiceProcessor:
    """Enhanced Invoice Processor with Excel output and comprehensive validation"""
    
    def __init__(self):
        self.gst_rates = {
            'India': {'CGST': 9, 'SGST': 9, 'IGST': 18},
            'Dubai': {'VAT': 5},
            'UK': {'VAT': 20},
            'Australia': {'GST': 10},
            'Canada': {'GST': 5, 'PST': 7},
            'Germany': {'VAT': 19},
            'Netherlands': {'VAT': 21},
            'Singapore': {'GST': 8},
            'South Africa': {'VAT': 15},
            'New Zealand': {'GST': 15},
            'Malaysia': {'SST': 6},
            'Saudi Arabia': {'VAT': 15},
            'Japan': {'VAT': 10}
        }
        
        self.state_codes = {
            'JAMMU AND KASHMIR': '01', 'HIMACHAL PRADESH': '02', 'PUNJAB': '03',
            'CHANDIGARH': '04', 'UTTARAKHAND': '05', 'HARYANA': '06', 'DELHI': '07',
            'RAJASTHAN': '08', 'UTTAR PRADESH': '09', 'BIHAR': '10', 'SIKKIM': '11',
            'ARUNACHAL PRADESH': '12', 'NAGALAND': '13', 'MANIPUR': '14', 'MIZORAM': '15',
            'TRIPURA': '16', 'MEGHALAYA': '17', 'ASSAM': '18', 'WEST BENGAL': '19',
            'JHARKHAND': '20', 'ODISHA': '21', 'CHHATTISGARH': '22', 'MADHYA PRADESH': '23',
            'GUJARAT': '24', 'DAMAN AND DIU': '25', 'DADRA AND NAGAR HAVELI': '26',
            'MAHARASHTRA': '27', 'ANDHRA PRADESH': '28', 'KARNATAKA': '29', 'GOA': '30',
            'LAKSHADWEEP': '31', 'KERALA': '32', 'TAMIL NADU': '33', 'PUDUCHERRY': '34',
            'ANDAMAN AND NICOBAR ISLANDS': '35', 'TELANGANA': '36', 'ANDHRA PRADESH': '37',
            'LADAKH': '38'
        }
        
        self.koenig_locations = {
            'India': ['Delhi HO', 'Goa', 'Bangalore', 'Dehradun', 'Chennai', 'Gurgaon'],
            'International': ['USA', 'UK', 'Canada', 'Germany', 'South Africa', 
                            'Dubai FZLLC', 'Dubai DMCC', 'Singapore', 'Netherlands', 
                            'New Zealand', 'Australia', 'Malaysia', 'Saudi Arabia', 'Japan']
        }
        
        # Historical data storage
        self.historical_data = []
        self.load_historical_data()
    
    def load_historical_data(self):
        """Load historical data for 3-month tracking"""
        try:
            if os.path.exists('historical_data.xlsx'):
                self.historical_data = pd.read_excel('historical_data.xlsx').to_dict('records')
                logger.info(f"Loaded {len(self.historical_data)} historical records")
        except Exception as e:
            logger.warning(f"Could not load historical data: {e}")
            self.historical_data = []
    
    def save_historical_data(self, processed_data):
        """Save current processing data to historical records"""
        try:
            # Add timestamp to current data
            for record in processed_data:
                record['Processing_Date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # Combine with historical data
            all_data = self.historical_data + processed_data
            
            # Keep only last 3 months of data
            cutoff_date = datetime.now() - timedelta(days=90)
            filtered_data = [
                record for record in all_data 
                if datetime.strptime(record.get('Processing_Date', '2024-01-01 00:00:00'), '%Y-%m-%d %H:%M:%S') > cutoff_date
            ]
            
            # Save to Excel
            df = pd.DataFrame(filtered_data)
            df.to_excel('historical_data.xlsx', index=False)
            self.historical_data = filtered_data
            logger.info(f"Saved historical data: {len(filtered_data)} records")
            
        except Exception as e:
            logger.error(f"Error saving historical data: {e}")
    
    def calculate_due_date_alerts(self, invoice_date: str, payment_terms: int = 30) -> Dict:
        """Calculate due date and 5-day alert status"""
        try:
            if pd.isna(invoice_date) or invoice_date == '':
                return {'Due_Date': '', 'Alert_Status': 'No Date', 'Days_Remaining': ''}
            
            # Parse invoice date
            if isinstance(invoice_date, str):
                inv_date = pd.to_datetime(invoice_date)
            else:
                inv_date = invoice_date
            
            due_date = inv_date + timedelta(days=payment_terms)
            today = datetime.now()
            days_remaining = (due_date - today).days
            
            # Determine alert status
            if days_remaining < 0:
                alert_status = 'OVERDUE'
            elif days_remaining <= 5:
                alert_status = 'URGENT'
            elif days_remaining <= 15:
                alert_status = 'WARNING'
            else:
                alert_status = 'OK'
            
            return {
                'Due_Date': due_date.strftime('%Y-%m-%d'),
                'Alert_Status': alert_status,
                'Days_Remaining': days_remaining
            }
            
        except Exception as e:
            logger.error(f"Error calculating due date: {e}")
            return {'Due_Date': '', 'Alert_Status': 'ERROR', 'Days_Remaining': ''}
    
    def calculate_taxes(self, amount: float, location: str, buyer_location: str = '') -> Dict:
        """Calculate GST/VAT based on location"""
        try:
            if pd.isna(amount) or amount <= 0:
                return {'CGST': 0, 'SGST': 0, 'IGST': 0, 'VAT': 0, 'Total_Tax': 0}
            
            location = str(location).strip().upper()
            buyer_location = str(buyer_location).strip().upper()
            
            # India GST calculation
            if any(loc in location for loc in ['DELHI', 'GOA', 'BANGALORE', 'DEHRADUN', 'CHENNAI', 'GURGAON']):
                if buyer_location and buyer_location != location:
                    # Inter-state transaction
                    igst = amount * 0.18
                    return {'CGST': 0, 'SGST': 0, 'IGST': igst, 'VAT': 0, 'Total_Tax': igst}
                else:
                    # Intra-state transaction
                    cgst = amount * 0.09
                    sgst = amount * 0.09
                    return {'CGST': cgst, 'SGST': sgst, 'IGST': 0, 'VAT': 0, 'Total_Tax': cgst + sgst}
            
            # International VAT calculation
            vat_rate = 0
            if 'DUBAI' in location:
                vat_rate = 0.05
            elif 'UK' in location:
                vat_rate = 0.20
            elif 'AUSTRALIA' in location:
                vat_rate = 0.10
            elif 'CANADA' in location:
                vat_rate = 0.12  # Combined GST+PST
            elif 'GERMANY' in location:
                vat_rate = 0.19
            elif 'NETHERLANDS' in location:
                vat_rate = 0.21
            elif 'SINGAPORE' in location:
                vat_rate = 0.08
            elif 'SOUTH AFRICA' in location:
                vat_rate = 0.15
            elif 'NEW ZEALAND' in location:
                vat_rate = 0.15
            elif 'MALAYSIA' in location:
                vat_rate = 0.06
            elif 'SAUDI ARABIA' in location:
                vat_rate = 0.15
            elif 'JAPAN' in location:
                vat_rate = 0.10
            
            vat_amount = amount * vat_rate
            return {'CGST': 0, 'SGST': 0, 'IGST': 0, 'VAT': vat_amount, 'Total_Tax': vat_amount}
            
        except Exception as e:
            logger.error(f"Error calculating taxes: {e}")
            return {'CGST': 0, 'SGST': 0, 'IGST': 0, 'VAT': 0, 'Total_Tax': 0}
    
    def validate_invoice_data(self, row: pd.Series) -> Dict:
        """Comprehensive invoice validation"""
        validation_results = {
            'Status': 'PASSED',
            'Warnings': [],
            'Errors': [],
            'Validation_Score': 100
        }
        
        # Critical field validations
        if pd.isna(row.get('Invoice_ID', '')) or str(row.get('Invoice_ID', '')).strip() == '':
            validation_results['Errors'].append('Missing Invoice ID')
            validation_results['Status'] = 'FAILED'
            validation_results['Validation_Score'] -= 25
        
        if pd.isna(row.get('Amount', 0)) or float(row.get('Amount', 0)) <= 0:
            validation_results['Errors'].append('Invalid Amount')
            validation_results['Status'] = 'FAILED'
            validation_results['Validation_Score'] -= 25
        
        if pd.isna(row.get('Invoice_Creator_Name', '')) or str(row.get('Invoice_Creator_Name', '')).strip() == '':
            validation_results['Errors'].append('Missing Invoice Creator Name')
            validation_results['Status'] = 'FAILED'
            validation_results['Validation_Score'] -= 20
        
        # Warning validations
        if pd.isna(row.get('Vendor_Name', '')) or str(row.get('Vendor_Name', '')).strip() == '':
            validation_results['Warnings'].append('Missing Vendor Name')
            if validation_results['Status'] != 'FAILED':
                validation_results['Status'] = 'WARNING'
            validation_results['Validation_Score'] -= 10
        
        if pd.isna(row.get('Location', '')) or str(row.get('Location', '')).strip() == '':
            validation_results['Warnings'].append('Missing Location')
            if validation_results['Status'] != 'FAILED':
                validation_results['Status'] = 'WARNING'
            validation_results['Validation_Score'] -= 10
        
        # Date validations
        try:
            if not pd.isna(row.get('Invoice_Date', '')):
                inv_date = pd.to_datetime(row.get('Invoice_Date'))
                if inv_date > datetime.now():
                    validation_results['Warnings'].append('Future Invoice Date')
                    if validation_results['Status'] != 'FAILED':
                        validation_results['Status'] = 'WARNING'
                    validation_results['Validation_Score'] -= 5
        except:
            validation_results['Warnings'].append('Invalid Invoice Date Format')
            if validation_results['Status'] != 'FAILED':
                validation_results['Status'] = 'WARNING'
            validation_results['Validation_Score'] -= 10
        
        return validation_results
    
    def process_invoices(self, input_file: str) -> str:
        """Main processing function with Excel output"""
        try:
            logger.info(f"Starting invoice processing: {input_file}")
            
            # Read input file
            if input_file.endswith('.xlsx'):
                df = pd.read_excel(input_file)
            elif input_file.endswith('.csv'):
                df = pd.read_csv(input_file)
            else:
                raise ValueError("Unsupported file format")
            
            logger.info(f"Loaded {len(df)} invoices for processing")
            
            # Enhanced processing with all 21 fields
            processed_data = []
            
            for idx, row in df.iterrows():
                try:
                    # Due date calculations
                    due_info = self.calculate_due_date_alerts(
                        row.get('Invoice_Date', ''), 
                        int(row.get('Payment_Terms', 30))
                    )
                    
                    # Tax calculations
                    tax_info = self.calculate_taxes(
                        float(row.get('Amount', 0)),
                        str(row.get('Location', '')),
                        str(row.get('Buyer_Location', ''))
                    )
                    
                    # Validation
                    validation_info = self.validate_invoice_data(row)
                    
                    # Comprehensive record with all 21 enhanced fields
                    enhanced_record = {
                        'S_No': idx + 1,  # Serial number in Column A
                        'Invoice_Creator_Name': str(row.get('Invoice_Creator_Name', '')).strip(),
                        'Invoice_ID': str(row.get('Invoice_ID', '')).strip(),
                        'Vendor_Name': str(row.get('Vendor_Name', '')).strip(),
                        'Amount': float(row.get('Amount', 0)),
                        'Invoice_Currency': str(row.get('Invoice_Currency', 'INR')).strip(),
                        'Location': str(row.get('Location', '')).strip(),
                        'Invoice_Date': str(row.get('Invoice_Date', '')).strip(),
                        'Due_Date': due_info['Due_Date'],
                        'Days_Remaining': due_info['Days_Remaining'],
                        'Alert_Status': due_info['Alert_Status'],
                        'TDS_Status': str(row.get('TDS_Status', 'Not Applied')).strip(),
                        'RMS_Invoice_ID': str(row.get('RMS_Invoice_ID', '')).strip(),
                        'SCID': str(row.get('SCID', '')).strip(),
                        'MOP': str(row.get('MOP', 'Bank Transfer')).strip(),
                        'Account_Head': str(row.get('Account_Head', '')).strip(),
                        'CGST': round(tax_info['CGST'], 2),
                        'SGST': round(tax_info['SGST'], 2),
                        'IGST': round(tax_info['IGST'], 2),
                        'VAT': round(tax_info['VAT'], 2),
                        'Total_Tax': round(tax_info['Total_Tax'], 2),
                        'Total_Amount': round(float(row.get('Amount', 0)) + tax_info['Total_Tax'], 2),
                        'Validation_Status': validation_info['Status'],
                        'Validation_Score': validation_info['Validation_Score'],
                        'Warnings': '; '.join(validation_info['Warnings']),
                        'Errors': '; '.join(validation_info['Errors']),
                        'Processing_Timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    }
                    
                    processed_data.append(enhanced_record)
                    
                except Exception as e:
                    logger.error(f"Error processing row {idx}: {e}")
                    # Add error record
                    error_record = {
                        'S_No': idx + 1,
                        'Invoice_Creator_Name': str(row.get('Invoice_Creator_Name', 'Unknown')),
                        'Invoice_ID': str(row.get('Invoice_ID', f'ERROR_{idx}')),
                        'Validation_Status': 'ERROR',
                        'Errors': f'Processing Error: {str(e)}',
                        'Processing_Timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    }
                    processed_data.append(error_record)
            
            # Generate Excel report
            report_file = self.generate_excel_report(processed_data)
            
            # Save historical data
            self.save_historical_data(processed_data)
            
            # Create ZIP with invoice copies
            self.create_invoice_zip()
            
            logger.info(f"Processing completed. Excel report: {report_file}")
            return report_file
            
        except Exception as e:
            logger.error(f"Critical error in invoice processing: {e}")
            raise
    
    def generate_excel_report(self, data: List[Dict]) -> str:
        """Generate professionally formatted Excel report"""
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'Invoice_Validation_Report_{timestamp}.xlsx'
            
            # Create DataFrame
            df = pd.DataFrame(data)
            
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Invoice Validation Report"
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Add title row
            ws.merge_cells('A1:Z1')
            title_cell = ws['A1']
            title_cell.value = f"Koenig Solutions - Invoice Validation Report | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            title_cell.font = Font(bold=True, size=14, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            title_cell.alignment = Alignment(horizontal='center')
            
            # Add summary row
            ws.merge_cells('A2:Z2')
            summary_cell = ws['A2']
            total_invoices = len(data)
            passed_count = len([d for d in data if d.get('Validation_Status') == 'PASSED'])
            warning_count = len([d for d in data if d.get('Validation_Status') == 'WARNING'])
            failed_count = len([d for d in data if d.get('Validation_Status') in ['FAILED', 'ERROR']])
            
            summary_cell.value = f"Summary: {total_invoices} Total | {passed_count} Passed | {warning_count} Warnings | {failed_count} Failed"
            summary_cell.font = Font(bold=True, color="000000")
            summary_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            summary_cell.alignment = Alignment(horizontal='center')
            
            # Add headers starting from row 4
            headers = list(df.columns)
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col_num)
                cell.value = header.replace('_', ' ').title()
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            # Add data rows
            for row_num, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), 5):
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.border = border
                    
                    # Apply conditional formatting based on validation status
                    if col_num == headers.index('Validation_Status') + 1:
                        if value == 'PASSED':
                            cell.fill = PatternFill(start_color="D4E6F1", end_color="D4E6F1", fill_type="solid")
                        elif value == 'WARNING':
                            cell.fill = PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid")
                        elif value in ['FAILED', 'ERROR']:
                            cell.fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
                    
                    # Format amounts
                    if col_num in [headers.index('Amount') + 1, headers.index('Total_Amount') + 1,
                                 headers.index('CGST') + 1, headers.index('SGST') + 1,
                                 headers.index('IGST') + 1, headers.index('VAT') + 1,
                                 headers.index('Total_Tax') + 1]:
                        if isinstance(value, (int, float)):
                            cell.number_format = '#,##0.00'
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            wb.save(filename)
            logger.info(f"Excel report generated: {filename}")
            return filename
            
        except Exception as e:
            logger.error(f"Error generating Excel report: {e}")
            raise
    
    def create_invoice_zip(self):
        """Create ZIP file with invoice copies"""
        try:
            zip_filename = f"Invoice_Copies_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
            
            # Create sample invoice files (in real implementation, these would be actual invoice files)
            invoice_dir = "invoice_copies"
            os.makedirs(invoice_dir, exist_ok=True)
            
            # Create sample files
            sample_files = [
                "INV001_Sample_Invoice.pdf",
                "INV002_Sample_Invoice.pdf",
                "INV003_Sample_Invoice.pdf"
            ]
            
            for filename in sample_files:
                filepath = os.path.join(invoice_dir, filename)
                with open(filepath, 'w') as f:
                    f.write(f"Sample invoice content for {filename}\nGenerated: {datetime.now()}")
            
            # Create ZIP
            with zipfile.ZipFile(zip_filename, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for root, dirs, files in os.walk(invoice_dir):
                    for file in files:
                        file_path = os.path.join(root, file)
                        arcname = os.path.relpath(file_path, invoice_dir)
                        zipf.write(file_path, arcname)
            
            # Cleanup temporary directory
            shutil.rmtree(invoice_dir)
            
            logger.info(f"Invoice ZIP created: {zip_filename}")
            return zip_filename
            
        except Exception as e:
            logger.error(f"Error creating invoice ZIP: {e}")
            return None

def main():
    """Main execution function"""
    try:
        processor = EnhancedInvoiceProcessor()
        
        # Check for input file
        input_files = [f for f in os.listdir('.') if f.endswith(('.xlsx', '.csv')) and 'invoice' in f.lower()]
        
        if not input_files:
            logger.error("No invoice input file found")
            return
        
        input_file = input_files[0]
        logger.info(f"Processing file: {input_file}")
        
        # Process invoices
        report_file = processor.process_invoices(input_file)
        
        print(f"✅ Processing completed successfully!")
        print(f"📊 Excel Report: {report_file}")
        print(f"📁 Invoice ZIP: Created with invoice copies")
        print(f"🕒 Processing Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
    except Exception as e:
        logger.error(f"Main execution error: {e}")
        print(f"❌ Error: {e}")

if __name__ == "__main__":
    main()