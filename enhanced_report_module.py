import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import sqlite3

class EnhancedInvoiceValidator:
    def __init__(self):
        self.setup_database()
        self.initialize_lookup_tables()
    
    def initialize_lookup_tables(self):
        """Initialize all GST/VAT rates and state codes"""
        self.gst_vat_rates = {
            'india': {
                'intra_state': {'CGST': 9, 'SGST': 9, 'total': 18},
                'inter_state': {'IGST': 18, 'total': 18}
            },
            'dubai': {'VAT': 5}, 'usa': {'VAT': 0}, 'uk': {'VAT': 20},
            'australia': {'GST': 10}, 'singapore': {'GST': 9},
            'south_africa': {'VAT': 15}, 'netherlands': {'VAT': 21},
            'canada': {'GST': 5, 'HST_ON': 13, 'HST_MARITIME': 15},
            'new_zealand': {'VAT': 15}, 'germany': {'VAT': 19},
            'saudi_arab': {'VAT': 15}, 'malaysia': {'SST': 8}
        }
        
        self.state_codes = {
            'Jammu & Kashmir': '01', 'Himachal Pradesh': '02', 'Punjab': '03',
            'Chandigarh': '04', 'Uttarakhand': '05', 'Haryana': '06',
            'Delhi': '07', 'Rajasthan': '08', 'Uttar Pradesh': '09',
            'Bihar': '10', 'Sikkim': '11', 'Arunachal Pradesh': '12',
            'Nagaland': '13', 'Manipur': '14', 'Mizoram': '15',
            'Tripura': '16', 'Meghalaya': '17', 'Assam': '18',
            'West Bengal': '19', 'Jharkhand': '20', 'Odisha': '21',
            'Chhattisgarh': '22', 'Madhya Pradesh': '23', 'Gujarat': '24',
            'Daman & Diu': '25', 'Dadra & Nagar Haveli': '26',
            'Maharashtra': '27', 'Andhra Pradesh (Old)': '28', 'Karnataka': '29',
            'Goa': '30', 'Lakshadweep': '31', 'Kerala': '32',
            'Tamil Nadu': '33', 'Puducherry': '34', 'Andaman & Nicobar Islands': '35',
            'Telangana': '36', 'Andhra Pradesh (Newly Added)': '37', 'Ladakh': '38'
        }
        
        self.koenig_branches = {
            'india': ['Delhi HO', 'Goa', 'Bangalore', 'Dehradun', 'Chennai', 'Gurgaon'],
            'international': ['USA', 'UK', 'Canada', 'Germany', 'South Africa', 
                             'Dubai FZLLC', 'Dubai DMCC', 'Singapore', 'Netherlands', 
                             'New Zealand', 'Australia', 'Malaysia', 'Saudi Arab', 'Japan']
        }
    
    def setup_database(self):
        """Setup historical tracking database"""
        conn = sqlite3.connect('invoice_history.db')
        cursor = conn.cursor()
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS invoice_snapshots (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                invoice_id TEXT,
                snapshot_data TEXT,
                snapshot_date DATETIME,
                validation_run_id TEXT
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS change_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                invoice_id TEXT,
                field_changed TEXT,
                old_value TEXT,
                new_value TEXT,
                change_date DATETIME,
                change_type TEXT
            )
        ''')
        
        conn.commit()
        conn.close()
    
    def determine_location_entity(self, invoice_data):
        """Determine location and entity (Koenig/Rayontara)"""
        # Extract location from RMS data or vendor info
        location_raw = str(invoice_data.get('location', '')).lower()
        vendor_name = str(invoice_data.get('Vendor_Name', '')).lower()
        
        # Check for India branches
        for branch in self.koenig_branches['india']:
            if branch.lower() in location_raw:
                entity = 'Rayontara' if 'rayontara' in vendor_name else 'Koenig'
                return f"{branch} - {entity}", 'india'
        
        # Check for international locations
        for country in self.koenig_branches['international']:
            if country.lower().replace(' ', '').replace('_', '') in location_raw.replace(' ', '').replace('_', ''):
                return f"{country} - Koenig", country.lower().replace(' ', '_')
        
        return "Unknown - Koenig", 'unknown'
    
    def calculate_gst_vat(self, amount, location_country, supplier_state=None, buyer_state=None):
        """Calculate GST/VAT based on location and rules"""
        if not amount or amount <= 0:
            return {'error': 'Invalid amount'}
        
        amount = float(amount)
        
        if location_country == 'india':
            # India GST calculation
            if supplier_state and buyer_state:
                if supplier_state == buyer_state:
                    # Intra-state: CGST + SGST
                    cgst = amount * 0.09
                    sgst = amount * 0.09
                    return {
                        'CGST_9%': round(cgst, 2),
                        'SGST_9%': round(sgst, 2),
                        'Total_Tax': round(cgst + sgst, 2),
                        'Tax_Type': 'Intra-State GST'
                    }
                else:
                    # Inter-state: IGST
                    igst = amount * 0.18
                    return {
                        'IGST_18%': round(igst, 2),
                        'Total_Tax': round(igst, 2),
                        'Tax_Type': 'Inter-State GST'
                    }
            else:
                # Default to IGST if state info missing
                igst = amount * 0.18
                return {
                    'IGST_18%': round(igst, 2),
                    'Total_Tax': round(igst, 2),
                    'Tax_Type': 'GST (State Info Missing)'
                }
        
        elif location_country in self.gst_vat_rates:
            # International VAT/GST
            rates = self.gst_vat_rates[location_country]
            tax_info = {}
            total_tax = 0
            
            for tax_type, rate in rates.items():
                if isinstance(rate, (int, float)) and rate > 0:
                    tax_amount = amount * (rate / 100)
                    tax_info[f'{tax_type}_{rate}%'] = round(tax_amount, 2)
                    total_tax += tax_amount
            
            tax_info['Total_Tax'] = round(total_tax, 2)
            tax_info['Tax_Type'] = f'{location_country.title()} VAT/GST'
            return tax_info
        
        return {'Tax_Type': 'Unknown Location', 'Total_Tax': 0}
    
    def calculate_due_date(self, invoice_date_str, payment_terms=30):
        """Calculate due date and check for alerts"""
        try:
            if isinstance(invoice_date_str, str):
                invoice_date = datetime.strptime(invoice_date_str, '%Y-%m-%d')
            else:
                invoice_date = invoice_date_str
            
            due_date = invoice_date + timedelta(days=int(payment_terms))
            days_remaining = (due_date - datetime.now()).days
            
            return {
                'Due_Date': due_date.strftime('%Y-%m-%d'),
                'Days_Remaining': days_remaining,
                'Alert_Status': 'URGENT' if days_remaining <= 5 else 'OK'
            }
        except:
            return {
                'Due_Date': 'Invalid Date',
                'Days_Remaining': 0,
                'Alert_Status': 'ERROR'
            }
    
    def track_historical_changes(self, current_data, validation_run_id):
        """Track changes from previous validation runs"""
        conn = sqlite3.connect('invoice_history.db')
        cursor = conn.cursor()
        
        changes_detected = []
        
        # Get previous snapshot for comparison
        cursor.execute('''
            SELECT invoice_id, snapshot_data 
            FROM invoice_snapshots 
            WHERE snapshot_date >= ? 
            ORDER BY snapshot_date DESC
        ''', (datetime.now() - timedelta(days=90),))  # 3 months
        
        previous_snapshots = {}
        for row in cursor.fetchall():
            if row[0] not in previous_snapshots:
                previous_snapshots[row[0]] = eval(row[1])  # Convert string back to dict
        
        # Compare current data with previous snapshots
        for current_invoice in current_data:
            invoice_id = current_invoice.get('Invoice_ID')
            if invoice_id in previous_snapshots:
                previous = previous_snapshots[invoice_id]
                
                for field, current_value in current_invoice.items():
                    previous_value = previous.get(field)
                    if str(current_value) != str(previous_value):
                        change_record = {
                            'invoice_id': invoice_id,
                            'field': field,
                            'old_value': previous_value,
                            'new_value': current_value,
                            'change_date': datetime.now()
                        }
                        changes_detected.append(change_record)
                        
                        # Log to database
                        cursor.execute('''
                            INSERT INTO change_log 
                            (invoice_id, field_changed, old_value, new_value, change_date, change_type)
                            VALUES (?, ?, ?, ?, ?, ?)
                        ''', (invoice_id, field, str(previous_value), str(current_value), 
                              datetime.now(), 'MODIFIED'))
        
        # Save current snapshot
        for invoice in current_data:
            cursor.execute('''
                INSERT INTO invoice_snapshots 
                (invoice_id, snapshot_data, snapshot_date, validation_run_id)
                VALUES (?, ?, ?, ?)
            ''', (invoice.get('Invoice_ID'), str(invoice), datetime.now(), validation_run_id))
        
        conn.commit()
        conn.close()
        
        return changes_detected
    
    def enhance_existing_data(self, original_df):
        """Enhance your existing DataFrame with all new fields"""
        enhanced_df = original_df.copy()
        
        # Add all new columns
        new_columns = {
            'Invoice_Currency': 'INR',  # Default
            'Location': '',
            'TDS_Status': 'Coming Soon',
            'RMS_Invoice_ID': '',
            'SCID': '',
            'MOP_Mode_of_Payment': '',
            'AH_Account_Head': '',
            'Due_Date': '',
            'Days_to_Due': 0,
            'Due_Date_Alert': 'OK',
            'Total_Invoice_Amount': 0,
            'Tax_Type': '',
            'Tax_Calculation_Status': 'Pending',
            'CGST_9%': 0,
            'SGST_9%': 0,
            'IGST_18%': 0,
            'VAT_Amount': 0,
            'Total_Tax_Calculated': 0,
            'Invoice_Uploaded_By': '',
            'Historical_Changes': 'None',
            'Validation_Run_ID': f"RUN_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        }
        
        for col, default_value in new_columns.items():
            enhanced_df[col] = default_value
        
        # Process each row for enhancements
        for idx, row in enhanced_df.iterrows():
            # Determine location and entity
            location, country = self.determine_location_entity(row)
            enhanced_df.at[idx, 'Location'] = location
            
            # Calculate GST/VAT
            amount = row.get('Amount', 0)
            if amount and amount > 0:
                tax_calc = self.calculate_gst_vat(amount, country)
                
                enhanced_df.at[idx, 'Tax_Type'] = tax_calc.get('Tax_Type', 'Unknown')
                enhanced_df.at[idx, 'Total_Tax_Calculated'] = tax_calc.get('Total_Tax', 0)
                enhanced_df.at[idx, 'Tax_Calculation_Status'] = 'Calculated'
                
                # Add specific tax amounts
                for tax_key, tax_amount in tax_calc.items():
                    if tax_key in enhanced_df.columns:
                        enhanced_df.at[idx, tax_key] = tax_amount
            
            # Calculate due dates
            invoice_date = row.get('Invoice_Date')
            if invoice_date:
                due_info = self.calculate_due_date(invoice_date)
                enhanced_df.at[idx, 'Due_Date'] = due_info['Due_Date']
                enhanced_df.at[idx, 'Days_to_Due'] = due_info['Days_Remaining']
                enhanced_df.at[idx, 'Due_Date_Alert'] = due_info['Alert_Status']
            
            # Mock RMS data (you'll replace with actual RMS API calls)
            enhanced_df.at[idx, 'RMS_Invoice_ID'] = f"RMS_{row.get('Invoice_ID', 'Unknown')}"
            enhanced_df.at[idx, 'SCID'] = f"SCID_{idx + 1:04d}"
            enhanced_df.at[idx, 'MOP_Mode_of_Payment'] = 'Bank Transfer'  # Default
            enhanced_df.at[idx, 'AH_Account_Head'] = 'General Expense'  # Default
            enhanced_df.at[idx, 'Invoice_Uploaded_By'] = 'system@koenig.com'  # Default
        
        return enhanced_df

class EnhancedEmailGenerator:
    def __init__(self):
        pass
    
    def generate_enhanced_email_body(self, validation_summary, enhanced_df, changes_detected):
        """Generate enhanced email maintaining your current format"""
        current_date = datetime.now().strftime('%Y-%m-%d')
        
        # Calculate enhanced statistics
        total_invoices = len(enhanced_df)
        passed_invoices = len(enhanced_df[enhanced_df['Validation_Status'] == 'Valid'])
        failed_invoices = len(enhanced_df[enhanced_df['Validation_Status'] == 'Invalid'])
        warnings = len(enhanced_df[enhanced_df['Validation_Status'] == 'Warning'])
        
        # Due date alerts
        urgent_dues = len(enhanced_df[enhanced_df['Due_Date_Alert'] == 'URGENT'])
        
        # Currency breakdown
        currencies = enhanced_df['Invoice_Currency'].value_counts().to_dict()
        
        # Location breakdown
        locations = enhanced_df['Location'].str.split(' -').str[0].value_counts().to_dict()
        
        email_body = f"""
        📊 Enhanced Invoice Validation Summary - {current_date}

        📅 Validation Period
        Current Batch: {validation_summary.get('current_batch_start', 'N/A')} to {validation_summary.get('current_batch_end', 'N/A')}
        Cumulative Range: {validation_summary.get('cumulative_start', 'N/A')} to {validation_summary.get('cumulative_end', 'N/A')}
        Total Coverage: {validation_summary.get('total_days', 'N/A')} days

        ✅ Total Invoices
        {total_invoices}

        ✅ Passed
        {passed_invoices} ({(passed_invoices/total_invoices*100):.1f}%)

        ⚠️ Warnings  
        {warnings}

        ❌ Failed
        {failed_invoices}

        🆕 ENHANCED FEATURES

        💱 Currency Breakdown
        """
        
        for currency, count in currencies.items():
            email_body += f"{currency}: {count} invoices\n        "
        
        email_body += f"""
        
        🌍 Location Analysis
        """
        
        for location, count in locations.items():
            email_body += f"{location}: {count} invoices\n        "
        
        email_body += f"""
        
        ⏰ Due Date Alerts
        Urgent (≤5 days): {urgent_dues} invoices
        
        🔄 Historical Changes Detected
        Modifications/Deletions: {len(changes_detected)} records
        
        💰 Tax Compliance Summary
        GST/VAT Calculated: {len(enhanced_df[enhanced_df['Tax_Calculation_Status'] == 'Calculated'])} invoices
        Multi-location Coverage: {len(locations)} locations
        """
        
        # Keep your existing issues analysis
        email_body += f"""
        
        🔍 Top Validation Issues
        Missing Invoice Creator Name: {len(enhanced_df[enhanced_df['Invoice_Creator_Name'].isna()])} invoices
        Missing GST Number: {len(enhanced_df[enhanced_df['GST_Number'].isna()])} invoices
        Tax Calculation Pending: {len(enhanced_df[enhanced_df['Tax_Calculation_Status'] == 'Pending'])} invoices
        Missing RMS Data: {len(enhanced_df[enhanced_df['SCID'].isna()])} invoices
        
        👤 Invoice Creator Analysis
        Total Creators: {enhanced_df['Invoice_Creator_Name'].nunique()}
        Unknown Creators: {len(enhanced_df[enhanced_df['Invoice_Creator_Name'].isna()])} invoices
        
        📈 Overall Health Score
        {'🟢 Good' if passed_invoices/total_invoices > 0.8 else '🟡 Fair' if passed_invoices/total_invoices > 0.5 else '🔴 Needs Attention'} - {(passed_invoices/total_invoices*100):.1f}% of invoices passed validation
        
        🆕 ENHANCED ATTACHMENTS
        📎 Attachments
        ✅ Enhanced Invoice Validation Report (Excel with {len(enhanced_df.columns)} fields)
        🗂️ Invoice Files from RMS (ZIP folder - validation period)
        📋 Historical Changes Log (CSV - 3 months tracking)  
        💰 GST/VAT Compliance Summary (PDF)
        🌍 Multi-location Tax Report (Excel)
        
        Additional Enhanced Information
        🔄 Historical Tracking: Active (3 months)
        💱 Multi-currency Support: {len(currencies)} currencies
        🌍 Global Coverage: {len(locations)} locations
        ⏰ Due Date Monitoring: {urgent_dues} urgent alerts
        """
        
        return email_body
    
    def create_enhanced_excel_report(self, enhanced_df, output_filename=None):
        """Create enhanced Excel report with formatting"""
        if not output_filename:
            timestamp = datetime.now().strftime('%Y-%m-%d')
            output_filename = f'enhanced_invoice_validation_detailed_{timestamp}.xlsx'
        
        with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
            # Main report
            enhanced_df.to_excel(writer, sheet_name='Enhanced Validation Report', index=False)
            
            # Summary sheet
            summary_data = {
                'Metric': ['Total Invoices', 'Valid', 'Invalid', 'Warnings', 'Currencies', 'Locations', 'Urgent Due Dates'],
                'Count': [
                    len(enhanced_df),
                    len(enhanced_df[enhanced_df['Validation_Status'] == 'Valid']),
                    len(enhanced_df[enhanced_df['Validation_Status'] == 'Invalid']),
                    len(enhanced_df[enhanced_df['Validation_Status'] == 'Warning']),
                    enhanced_df['Invoice_Currency'].nunique(),
                    enhanced_df['Location'].str.split(' -').str[0].nunique(),
                    len(enhanced_df[enhanced_df['Due_Date_Alert'] == 'URGENT'])
                ]
            }
            pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', index=False)
            
            # Tax breakdown sheet
            tax_summary = enhanced_df.groupby(['Location', 'Tax_Type'])[['Total_Tax_Calculated']].sum().reset_index()
            tax_summary.to_excel(writer, sheet_name='Tax Breakdown', index=False)
            
            # Apply formatting
            workbook = writer.book
            
            # Format main sheet
            worksheet = writer.sheets['Enhanced Validation Report']
            
            # Header formatting
            for cell in worksheet[1]:
                cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                cell.font = Font(color='FFFFFF', bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Highlight urgent due dates
            for row in range(2, len(enhanced_df) + 2):
                if worksheet.cell(row=row, column=enhanced_df.columns.get_loc('Due_Date_Alert')+1).value == 'URGENT':
                    for col in range(1, len(enhanced_df.columns) + 1):
                        worksheet.cell(row=row, column=col).fill = PatternFill(
                            start_color='FFE6E6', end_color='FFE6E6', fill_type='solid'
                        )
            
            # Adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        return output_filename

# Integration with your existing system
def integrate_enhancements(existing_excel_path):
    """Main function to integrate all enhancements with your existing system"""
    
    # Initialize enhanced validator
    validator = EnhancedInvoiceValidator()
    email_generator = EnhancedEmailGenerator()
    
    # Load your existing data
    existing_df = pd.read_excel(existing_excel_path)
    
    # Generate validation run ID
    validation_run_id = f"RUN_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    
    # Track historical changes
    current_data = existing_df.to_dict('records')
    changes_detected = validator.track_historical_changes(current_data, validation_run_id)
    
    # Enhance the existing data
    enhanced_df = validator.enhance_existing_data(existing_df)
    
    # Create enhanced Excel report
    enhanced_excel_file = email_generator.create_enhanced_excel_report(enhanced_df)
    
    # Prepare validation summary (you'll need to adjust based on your data)
    validation_summary = {
        'current_batch_start': '2025-08-05',
        'current_batch_end': '2025-08-08', 
        'cumulative_start': '2025-07-18',
        'cumulative_end': '2025-08-08',
        'total_days': 22
    }
    
    # Generate enhanced email body
    enhanced_email_body = email_generator.generate_enhanced_email_body(
        validation_summary, enhanced_df, changes_detected
    )
    
    print("🚀 Enhancement Complete!")
    print(f"📊 Enhanced Excel Report: {enhanced_excel_file}")
    print(f"🔄 Historical Changes Detected: {len(changes_detected)}")
    print(f"💱 Currencies Processed: {enhanced_df['Invoice_Currency'].nunique()}")
    print(f"🌍 Locations Covered: {enhanced_df['Location'].str.split(' -').str[0].nunique()}")
    
    return enhanced_excel_file, enhanced_email_body, changes_detected

# Example usage
if __name__ == "__main__":
    # Replace with your actual Excel file path
    existing_file = "invoice_validation_detailed_2025-08-09.xlsx"
    
    enhanced_file, email_body, changes = integrate_enhancements(existing_file)
    
    print("\n" + "="*50)
    print("ENHANCED EMAIL BODY:")
    print("="*50)
    print(email_body)
