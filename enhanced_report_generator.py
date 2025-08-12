# enhanced_report_generator.py
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import zipfile
import os

class EnhancedReportGenerator:
    def __init__(self):
        self.gst_rates = {
            # India GST Logic
            'india': {'intra_state': {'CGST': 9, 'SGST': 9}, 'inter_state': {'IGST': 18}},
            # International VAT/GST rates
            'dubai': {'VAT': 5},
            'usa': {'VAT': 0},  # NA
            'uk': {'VAT': 20},
            'australia': {'GST': 10},
            'singapore': {'GST': 9},
            'south_africa': {'VAT': 15},
            'netherlands': {'VAT': 21},
            'canada': {'GST': 5, 'HST_ON': 13, 'HST_NB_NL_NS_PE': 15},
            'new_zealand': {'VAT': 15},
            'germany': {'VAT': 19},
            'saudi_arab': {'VAT': 15},
            'malaysia': {'SST': 8}
        }
        
        self.state_codes = {
            'Jammu & Kashmir': '01', 'Himachal Pradesh': '02', 'Punjab': '03',
            'Chandigarh': '04', 'Uttarakhand': '05', 'Haryana': '06',
            'Delhi': '07', 'Rajasthan': '08', 'Uttar Pradesh': '09',
            'Bihar': '10', 'Sikkim': '11', 'Arunachal Pradesh': '12',
            'Nagaland': '13', 'Manipur': '14', 'Mizoram': '15',
            'Tripura': '16', 'Meghalaya': '17', 'Assam': '18',
            'West Bengal': '19', 'Jharkhand': '20', 'Odisha': '21',
            'Chhattisgarh': '22', 'Madhya Pradesh': '23', 'Gujarat': '24',
            'Daman & Diu': '25', 'Dadra & Nagar Haveli': '26',
            'Maharashtra': '27', 'Andhra Pradesh (Old)': '28', 'Karnataka': '29',
            'Goa': '30', 'Lakshadweep': '31', 'Kerala': '32',
            'Tamil Nadu': '33', 'Puducherry': '34', 'Andaman & Nicobar Islands': '35',
            'Telangana': '36', 'Andhra Pradesh (Newly Added)': '37', 'Ladakh (Newly Added)': '38'
        }
        
        self.branch_locations = {
            'india': ['Delhi HO', 'Goa', 'Bangalore', 'Dehradun', 'Chennai', 'Gurgaon'],
            'international': ['USA', 'UK', 'Canada', 'Germany', 'South Africa', 
                             'Dubai (FZLLC)', 'Dubai (DMCC)', 'Singapore', 
                             'Netherlands', 'New Zealand', 'Australia', 
                             'Malaysia', 'Saudi Arab', 'Japan']
        }

    def determine_location_and_entity(self, invoice_data):
        """Determine location and entity (Koenig/Rayontara) from invoice data"""
        location = invoice_data.get('location', '').lower()
        entity = 'Koenig'  # Default
        
        # Check if it's India location
        if any(branch.lower() in location for branch in self.branch_locations['india']):
            if 'rayontara' in invoice_data.get('vendor_name', '').lower():
                entity = 'Rayontara'
            return f"{location.title()} - {entity}", True
        else:
            # International location
            country = next((country for country in self.branch_locations['international'] 
                          if country.lower() in location), 'Unknown')
            return f"{country} - {entity}", False

    def calculate_gst_vat(self, invoice_data, is_india, location):
        """Calculate GST/VAT based on location and invoice details"""
        total_amount = float(invoice_data.get('total_amount', 0))
        tax_details = {}
        
        if is_india:
            # India GST calculation
            supplier_state = invoice_data.get('supplier_state_code', '')
            buyer_state = invoice_data.get('buyer_state_code', '')
            
            if supplier_state == buyer_state:
                # Intra-state (CGST + SGST)
                cgst = total_amount * 0.09
                sgst = total_amount * 0.09
                tax_details = {
                    'CGST': {'rate': '9%', 'amount': cgst},
                    'SGST': {'rate': '9%', 'amount': sgst},
                    'total_tax': cgst + sgst
                }
            else:
                # Inter-state (IGST)
                igst = total_amount * 0.18
                tax_details = {
                    'IGST': {'rate': '18%', 'amount': igst},
                    'total_tax': igst
                }
        else:
            # International VAT/GST calculation
            country = location.split(' -')[0].lower().replace(' ', '_')
            if country in self.gst_rates:
                rates = self.gst_rates[country]
                for tax_type, rate in rates.items():
                    if rate > 0:
                        tax_amount = total_amount * (rate / 100)
                        tax_details[tax_type] = {'rate': f'{rate}%', 'amount': tax_amount}
                        tax_details['total_tax'] = tax_details.get('total_tax', 0) + tax_amount
        
        return tax_details

    def check_due_date_notification(self, invoice_date, payment_terms):
        """Calculate due date and check if notification is needed"""
        if not invoice_date or not payment_terms:
            return None, False
        
        due_date = invoice_date + timedelta(days=int(payment_terms))
        days_to_due = (due_date - datetime.now()).days
        
        notification_needed = days_to_due <= 5  # Notify if due within 5 days
        return due_date, notification_needed

    def generate_enhanced_report(self, invoice_data_list, validation_results):
        """Generate enhanced Excel report with all new fields"""
        enhanced_data = []
        
        for idx, invoice in enumerate(invoice_data_list):
            location, is_india = self.determine_location_and_entity(invoice)
            tax_details = self.calculate_gst_vat(invoice, is_india, location)
            due_date, notification_needed = self.check_due_date_notification(
                invoice.get('invoice_date'), invoice.get('payment_terms')
            )
            
            enhanced_record = {
                # Existing fields
                'Invoice Number': invoice.get('invoice_number'),
                'Vendor Name': invoice.get('vendor_name'),
                'Invoice Date': invoice.get('invoice_date'),
                'Validation Status': validation_results[idx].get('status', 'Pending'),
                'Issues Found': ', '.join(validation_results[idx].get('issues', [])),
                
                # NEW ENHANCED FIELDS
                'Invoice Currency': invoice.get('currency', 'INR'),
                'Location': location,
                'TDS Status': 'Coming Soon',  # Placeholder for TDS implementation
                'Invoice ID (RMS)': invoice.get('rms_invoice_id'),
                'SCID': invoice.get('scid'),
                'MOP (Mode of Payment)': invoice.get('mode_of_payment'),
                'AH (Account Head)': invoice.get('account_head'),
                'Due Date': due_date.strftime('%Y-%m-%d') if due_date else 'N/A',
                'Due Date Notification': 'YES' if notification_needed else 'NO',
                'Total Invoice Amount': invoice.get('total_amount'),
                'Invoice Uploaded By': invoice.get('uploaded_by'),
                
                # TAX BREAKDOWN
                'Tax Type': ', '.join(tax_details.keys()) if tax_details else 'N/A',
                'Tax Calculation Status': 'Verified' if tax_details else 'Pending',
                'Total Tax Amount': tax_details.get('total_tax', 0),
            }
            
            # Add individual tax components
            for tax_type, details in tax_details.items():
                if tax_type != 'total_tax':
                    enhanced_record[f'{tax_type} Rate'] = details['rate']
                    enhanced_record[f'{tax_type} Amount'] = details['amount']
            
            enhanced_data.append(enhanced_record)
        
        # Create DataFrame and Excel file
        df = pd.DataFrame(enhanced_data)
        
        # Apply conditional formatting
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'Enhanced_Invoice_Validation_Report_{timestamp}.xlsx'
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Validation Report', index=False)
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Validation Report']
            
            # Apply formatting
            from openpyxl.styles import PatternFill, Font
            
            # Header formatting
            for cell in worksheet[1]:
                cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                cell.font = Font(color='FFFFFF', bold=True)
            
            # Highlight due date notifications
            for row in range(2, len(enhanced_data) + 2):
                if worksheet[f'L{row}'].value == 'YES':  # Due Date Notification column
                    for col in range(1, len(enhanced_data[0]) + 1):
                        worksheet.cell(row=row, column=col).fill = PatternFill(
                            start_color='FFE6E6', end_color='FFE6E6', fill_type='solid'
                        )
        
        return filename

class HistoricalDataTracker:
    def __init__(self, db_path='invoice_history.db'):
        self.db_path = db_path
        self.setup_database()
    
    def setup_database(self):
        """Setup SQLite database for tracking historical changes"""
        import sqlite3
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS invoice_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                invoice_id TEXT,
                field_name TEXT,
                old_value TEXT,
                new_value TEXT,
                change_date DATETIME,
                change_type TEXT
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS invoice_snapshots (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                invoice_id TEXT,
                snapshot_data TEXT,
                snapshot_date DATETIME
            )
        ''')
        
        conn.commit()
        conn.close()
    
    def track_changes(self, current_invoices, previous_invoices):
        """Track changes in invoice data over past 3 months"""
        import sqlite3
        import json
        
        changes_detected = []
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        for current in current_invoices:
            invoice_id = current.get('invoice_id')
            previous = next((p for p in previous_invoices if p.get('invoice_id') == invoice_id), None)
            
            if previous:
                # Check for modifications
                for field, current_value in current.items():
                    previous_value = previous.get(field)
                    if current_value != previous_value:
                        change_record = {
                            'invoice_id': invoice_id,
                            'field_name': field,
                            'old_value': str(previous_value),
                            'new_value': str(current_value),
                            'change_date': datetime.now(),
                            'change_type': 'MODIFIED'
                        }
                        changes_detected.append(change_record)
                        
                        cursor.execute('''
                            INSERT INTO invoice_history 
                            (invoice_id, field_name, old_value, new_value, change_date, change_type)
                            VALUES (?, ?, ?, ?, ?, ?)
                        ''', (invoice_id, field, str(previous_value), str(current_value), 
                              datetime.now(), 'MODIFIED'))
        
        # Check for deletions
        current_ids = {inv.get('invoice_id') for inv in current_invoices}
        previous_ids = {inv.get('invoice_id') for inv in previous_invoices}
        deleted_ids = previous_ids - current_ids
        
        for deleted_id in deleted_ids:
            deleted_invoice = next(p for p in previous_invoices if p.get('invoice_id') == deleted_id)
            change_record = {
                'invoice_id': deleted_id,
                'field_name': 'FULL_RECORD',
                'old_value': json.dumps(deleted_invoice),
                'new_value': 'DELETED',
                'change_date': datetime.now(),
                'change_type': 'DELETED'
            }
            changes_detected.append(change_record)
            
            cursor.execute('''
                INSERT INTO invoice_history 
                (invoice_id, field_name, old_value, new_value, change_date, change_type)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (deleted_id, 'FULL_RECORD', json.dumps(deleted_invoice), 
                  'DELETED', datetime.now(), 'DELETED'))
        
        conn.commit()
        conn.close()
        
        return changes_detected
