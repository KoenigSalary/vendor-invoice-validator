# reporter.py

import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.filters import AutoFilter
from openpyxl.chart import BarChart, PieChart, Reference
from datetime import datetime
import logging

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def save_snapshot_report(data, start_date, end_date, output_dir="snapshots"):
    """
    Enhanced version of snapshot report with better error handling,
    statistics, and formatting
    """
    try:
        if not data:
            logger.warning("⚠️ No data to save in report.")
            return None

        # Ensure output directory exists
        os.makedirs(output_dir, exist_ok=True)

        # Create filename with timestamp
        timestamp = datetime.now().strftime("%H%M%S")
        filename = f"InvoiceSnapshot_{start_date.strftime('%Y-%m-%d')}_to_{end_date.strftime('%Y-%m-%d')}_{timestamp}.xlsx"
        filepath = os.path.join(output_dir, filename)

        # Convert data to DataFrame
        if isinstance(data, list):
            df = pd.DataFrame(data)
        elif isinstance(data, pd.DataFrame):
            df = data.copy()
        else:
            logger.error("❌ Invalid data format. Expected list or DataFrame.")
            return None

        if df.empty:
            logger.warning("⚠️ DataFrame is empty.")
            return None

        logger.info(f"📊 Creating snapshot report with {len(df)} records")

        # Standardize status column
        if 'Status' not in df.columns:
            # Try to derive status from other columns
            if 'Issues_Found' in df.columns:
                df['Status'] = df['Issues_Found'].apply(
                    lambda x: 'VALID' if not x or x == '' else 'FLAGGED'
                )
            elif 'Validation_Status' in df.columns:
                df['Status'] = df['Validation_Status'].str.upper()
            else:
                df['Status'] = 'UNKNOWN'

        # Ensure status values are standardized
        df['Status'] = df['Status'].fillna('UNKNOWN').str.upper()
        
        # Map common status variations
        status_mapping = {
            'VALID': 'VALID',
            'INVALID': 'INVALID', 
            'FLAGGED': 'FLAGGED',
            'NEW': 'FLAGGED',
            'ISSUES FOUND': 'FLAGGED',
            'UNKNOWN': 'UNKNOWN',
            'ERROR': 'INVALID'
        }
        df['Status'] = df['Status'].map(status_mapping).fillna('UNKNOWN')

        # Split by status
        valid_df = df[df["Status"] == "VALID"]
        invalid_df = df[df["Status"] == "INVALID"] 
        flagged_df = df[df["Status"] == "FLAGGED"]
        unknown_df = df[df["Status"] == "UNKNOWN"]

        # Create summary statistics
        summary_data = create_summary_statistics(df, start_date, end_date)

        # Write to Excel with pandas
        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            # Summary sheet first
            summary_df = pd.DataFrame(list(summary_data.items()), columns=['Metric', 'Value'])
            summary_df.to_excel(writer, sheet_name="SUMMARY", index=False)
            
            # Data sheets
            if not valid_df.empty:
                valid_df.to_excel(writer, sheet_name="VALID", index=False)
            if not invalid_df.empty:
                invalid_df.to_excel(writer, sheet_name="INVALID", index=False)
            if not flagged_df.empty:
                flagged_df.to_excel(writer, sheet_name="FLAGGED", index=False)
            if not unknown_df.empty:
                unknown_df.to_excel(writer, sheet_name="UNKNOWN", index=False)
            
            # All data sheet
            df.to_excel(writer, sheet_name="ALL_DATA", index=False)

        # Apply advanced formatting
        format_excel_report(filepath, summary_data)

        logger.info(f"✅ Snapshot report saved: {filepath}")
        return filepath

    except Exception as e:
        logger.error(f"❌ Failed to create snapshot report: {str(e)}")
        return None

def create_summary_statistics(df, start_date, end_date):
    """Create comprehensive summary statistics - FIXED VERSION"""
    try:
        total_records = len(df)
        status_counts = df['Status'].value_counts()
        
        # Calculate percentages
        valid_count = status_counts.get('VALID', 0)
        invalid_count = status_counts.get('INVALID', 0)
        flagged_count = status_counts.get('FLAGGED', 0)
        unknown_count = status_counts.get('UNKNOWN', 0)
        
        # Amount statistics (if available)
        amount_stats = {}
        if 'Amount' in df.columns:
            df['Amount_numeric'] = pd.to_numeric(df['Amount'], errors='coerce')
            amount_stats = {
                'Total Amount': f"₹{df['Amount_numeric'].sum():,.2f}",
                'Average Amount': f"₹{df['Amount_numeric'].mean():,.2f}",
                'Median Amount': f"₹{df['Amount_numeric'].median():,.2f}",
                'Max Amount': f"₹{df['Amount_numeric'].max():,.2f}",
                'Min Amount': f"₹{df['Amount_numeric'].min():,.2f}"
            }
        
        # Vendor statistics (if available)
        vendor_stats = {}
        if 'Vendor' in df.columns or 'PartyName' in df.columns:
            vendor_col = 'Vendor' if 'Vendor' in df.columns else 'PartyName'
            unique_vendors = df[vendor_col].nunique()
            vendor_stats = {
                'Unique Vendors': unique_vendors,
                'Top Vendor': df[vendor_col].mode().iloc[0] if not df[vendor_col].empty else 'N/A'
            }
        
        # FIXED: Build summary dictionary properly
        summary = {
            '=== REPORT OVERVIEW ===': '',
            'Report Generated': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'Period From': start_date.strftime('%Y-%m-%d'),
            'Period To': end_date.strftime('%Y-%m-%d'),
            'Total Records': total_records,
            'Overview_Separator': '',  # Using descriptive keys instead of empty strings
            '=== VALIDATION SUMMARY ===': '',
            'Valid Records': f"{valid_count} ({valid_count/total_records*100:.1f}%)" if total_records > 0 else "0 (0%)",
            'Invalid Records': f"{invalid_count} ({invalid_count/total_records*100:.1f}%)" if total_records > 0 else "0 (0%)",
            'Flagged Records': f"{flagged_count} ({flagged_count/total_records*100:.1f}%)" if total_records > 0 else "0 (0%)",
            'Unknown Status': f"{unknown_count} ({unknown_count/total_records*100:.1f}%)" if total_records > 0 else "0 (0%)",
            'Success Rate': f"{valid_count/total_records*100:.1f}%" if total_records > 0 else "0%"
        }
        
        # Add amount statistics if available
        if amount_stats:
            summary['Amount_Separator'] = ''
            summary['=== AMOUNT ANALYSIS ==='] = ''
            summary.update(amount_stats)
        
        # Add vendor statistics if available
        if vendor_stats:
            summary['Vendor_Separator'] = ''
            summary['=== VENDOR ANALYSIS ==='] = ''
            summary.update(vendor_stats)
        
        return summary
        
    except Exception as e:
        logger.error(f"❌ Error creating summary statistics: {str(e)}")
        return {'Error': 'Failed to generate statistics'}

def format_excel_report(filepath, summary_data):
    """Apply advanced formatting to the Excel report"""
    try:
        wb = load_workbook(filepath)
        
        # Color scheme
        color_map = {
            "VALID": "C6EFCE",      # Light Green
            "INVALID": "FFC7CE",    # Light Red
            "FLAGGED": "FFEB9C",    # Light Orange
            "UNKNOWN": "E1D5E7",    # Light Purple
            "SUMMARY": "D9E1F2"     # Light Blue
        }
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Format each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            if sheet_name == "SUMMARY":
                format_summary_sheet(ws, summary_data, color_map["SUMMARY"])
            else:
                format_data_sheet(ws, sheet_name, color_map, header_font, header_fill, border)
        
        wb.save(filepath)
        logger.info("✅ Excel formatting applied successfully")
        
    except Exception as e:
        logger.error(f"❌ Error formatting Excel report: {str(e)}")

def format_summary_sheet(ws, summary_data, bg_color):
    """Format the summary sheet with special styling"""
    try:
        # Apply background color to summary data
        summary_fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        
        # Style headers and important sections
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and str(cell.value).startswith('==='):
                    # Section headers
                    cell.font = Font(bold=True, size=14, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                elif cell.column == 1:  # Metric names
                    cell.font = Font(bold=True)
                    cell.fill = summary_fill
                else:  # Values
                    cell.fill = summary_fill
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
            
    except Exception as e:
        logger.error(f"❌ Error formatting summary sheet: {str(e)}")

def format_data_sheet(ws, sheet_name, color_map, header_font, header_fill, border):
    """Format data sheets with conditional formatting and filters"""
    try:
        if ws.max_row <= 1:  # Empty sheet
            return
            
        # Get the appropriate color for this sheet
        status_color = color_map.get(sheet_name, "FFFFFF")
        status_fill = PatternFill(start_color=status_color, end_color=status_color, fill_type="solid")
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Add filters
        if ws.max_row > 1:
            ws.auto_filter.ref = ws.dimensions
        
        # Format all cells
        for row_num, row in enumerate(ws.iter_rows(), 1):
            for cell in row:
                # Header row formatting
                if row_num == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    # Data row formatting
                    cell.fill = status_fill
                    
                # Apply border to all cells
                cell.border = border
        
        # Auto-size columns with reasonable limits
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)  # Max width of 30
            ws.column_dimensions[column_letter].width = adjusted_width
            
    except Exception as e:
        logger.error(f"❌ Error formatting data sheet {sheet_name}: {str(e)}")

def create_monthly_trend_report(data_folder="data", output_file="monthly_trend_report.xlsx"):
    """
    Create a trend report analyzing validation results over multiple months
    """
    try:
        logger.info("📈 Creating monthly trend report...")
        
        # Find all delta reports
        report_files = []
        for root, dirs, files in os.walk(data_folder):
            for file in files:
                if file.startswith("delta_report_") and file.endswith(".xlsx"):
                    report_files.append(os.path.join(root, file))
        
        if not report_files:
            logger.warning("⚠️ No delta reports found for trend analysis")
            return None
        
        # Process each report
        monthly_data = []
        for report_file in sorted(report_files):
            try:
                df = pd.read_excel(report_file)
                
                # Extract date from filename
                filename = os.path.basename(report_file)
                date_str = filename.replace("delta_report_", "").replace(".xlsx", "")
                report_date = datetime.strptime(date_str, "%Y-%m-%d")
                
                # Calculate metrics
                total_records = len(df)
                if total_records > 0:
                    status_counts = df.get('Status', pd.Series()).value_counts()
                    
                    monthly_data.append({
                        'Date': report_date,
                        'Month': report_date.strftime('%Y-%m'),
                        'Total_Records': total_records,
                        'Valid_Records': status_counts.get('VALID', status_counts.get('Valid', 0)),
                        'Invalid_Records': status_counts.get('INVALID', status_counts.get('Flagged', 0)),
                        'Success_Rate': status_counts.get('VALID', status_counts.get('Valid', 0)) / total_records * 100
                    })
                    
            except Exception as e:
                logger.warning(f"⚠️ Could not process {report_file}: {str(e)}")
                continue
        
        if not monthly_data:
            logger.warning("⚠️ No valid data found for trend report")
            return None
        
        # Create trend DataFrame
        trend_df = pd.DataFrame(monthly_data)
        
        # Save trend report
        output_path = os.path.join("reports", output_file)
        os.makedirs("reports", exist_ok=True)
        
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            trend_df.to_excel(writer, sheet_name="Monthly_Trends", index=False)
        
        logger.info(f"✅ Monthly trend report created: {output_path}")
        return output_path
        
    except Exception as e:
        logger.error(f"❌ Failed to create trend report: {str(e)}")
        return None

# For backward compatibility
def save_snapshot_report_simple(data, start_date, end_date):
    """Simple version that maintains original function signature"""
    return save_snapshot_report(data, start_date, end_date)

# Test function
def test_reporter():
    """Test the reporter with sample data"""
    try:
        logger.info("🧪 Testing reporter with sample data...")
        
        # Create sample data
        sample_data = [
            {
                'Invoice_No': 'INV001',
                'Vendor': 'ABC Corp',
                'Amount': 10000,
                'Status': 'VALID',
                'Issues': ''
            },
            {
                'Invoice_No': 'INV002', 
                'Vendor': 'XYZ Ltd',
                'Amount': 15000,
                'Status': 'FLAGGED',
                'Issues': 'Missing GSTIN'
            },
            {
                'Invoice_No': 'INV003',
                'Vendor': 'DEF Inc',
                'Amount': 8000,
                'Status': 'INVALID',
                'Issues': 'Invalid amount format'
            }
        ]
        
        start_date = datetime(2025, 7, 17)
        end_date = datetime(2025, 7, 20)
        
        result = save_snapshot_report(sample_data, start_date, end_date, "test_reports")
        
        if result:
            logger.info(f"✅ Test successful! Report saved: {result}")
            return True
        else:
            logger.error("❌ Test failed!")
            return False
            
    except Exception as e:
        logger.error(f"❌ Test error: {str(e)}")
        return False

if __name__ == "__main__":
    # Run test
    test_reporter()
    
    # Create trend report if data exists
    create_monthly_trend_report()
